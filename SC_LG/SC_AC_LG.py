"""
ETAP Short-Circuit LG Report — Extractor de resultados SLG
===========================================================

Este script lee el archivo .xls que genera ETAP para el reporte de
cortocircuito línea-a-tierra (SLG) y produce una tabla resumen en Excel
con el formato de la "SLG short circuit current results".

La tabla de salida contiene:
    Bus name  |  Bus (kV)  |  3I0 Symm Current 1/2 cycle [kA]

Uso desde la terminal:
    python etap_slg_extractor.py mi_reporte.xls
    python etap_slg_extractor.py mi_reporte.xls --output resultado.xlsx
    python etap_slg_extractor.py mi_reporte.xls --title "SLG short circuit current results"
    python etap_slg_extractor.py mi_reporte.xls --mw 2.8

Requisitos (instalar una sola vez):
    pip install pandas xlrd openpyxl

Columnas clave del archivo ETAP (posición base-0):
    col  1  → From Bus ID  (nombre del bus faultado, en la fila "Total")
    col  5  → To Bus ID    (vale "Total" en la fila principal de cada bloque)
    col 11  → Prefault kV  (en la fila "Prefault voltage...")
    col 28  → Ia Mag [kA]  (corriente de fase A = 3·I0 para falla SLG)
    col 42  → I1 [kA]      (corriente de secuencia positiva)
    col 44  → I2 [kA]      (corriente de secuencia negativa)
    col 45  → I0 [kA]      (corriente de secuencia cero)

En una falla SLG: Ia = 3·I0 = I1 + I2 + I0
El "3I0 Symm Current 1/2 cycle" de la tabla = Ia (columna 28) del bloque Total.
"""

import argparse
import re
import sys
from pathlib import Path

import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


# ──────────────────────────────────────────────────────────────────────────────
# PASO 1 — Leer el archivo XLS
# ──────────────────────────────────────────────────────────────────────────────
def read_report(filepath: str) -> pd.DataFrame:
    """
    Carga el archivo ETAP sin cabeceras (header=None) porque el reporte
    tiene múltiples filas de encabezado y bloques de datos entremezclados.

    .xls  → motor xlrd  (formato Excel 97-2003, el que genera ETAP)
    .xlsx → motor openpyxl (formato moderno)
    """
    path = Path(filepath)
    if not path.exists():
        sys.exit(f"Archivo no encontrado: {filepath}")

    engine = "xlrd" if path.suffix.lower() == ".xls" else "openpyxl"
    print(f"Leyendo: {path.name}  (motor: {engine})")

    df = pd.read_excel(filepath, engine=engine, header=None)
    print(f"   → {len(df)} filas × {len(df.columns)} columnas")
    return df


# ──────────────────────────────────────────────────────────────────────────────
# PASO 2 — Extraer el voltaje prefalla de un bloque
# ──────────────────────────────────────────────────────────────────────────────
def extract_prefault_kv(df: pd.DataFrame, fault_row: int) -> float:
    """
    Busca la fila "Prefault voltage = X.XXX kV" que aparece justo
    después de la línea "Fault at bus: ...".
    El valor en kV está en el texto de la columna 1 de esa fila.
    """
    for i in range(fault_row, fault_row + 5):
        text = str(df.iloc[i][1])
        match = re.search(r"Prefault voltage\s*=\s*([\d.]+)\s*kV", text)
        if match:
            return float(match.group(1))
    return float("nan")


# ──────────────────────────────────────────────────────────────────────────────
# PASO 3 — Extraer todos los bloques de falla
# ──────────────────────────────────────────────────────────────────────────────
def extract_slg_results(df: pd.DataFrame) -> list[dict]:
    """
    Recorre el DataFrame buscando líneas "Fault at bus: <nombre>".
    Para cada bloque, localiza la fila con To Bus == "Total" y extrae:
        - Bus name: nombre del bus faultado
        - Bus kV:   voltaje nominal del bus (de la prefault line)
        - Ia [kA]:  corriente total de falla = 3·I0 (columna 28)
        - I1 [kA]:  secuencia positiva (columna 42)

    ¿Por qué col 28 es la corriente total?
        En una falla SLG, la corriente de falla es:
            Ia = 3 · I0 = I1 + I2 + I0
        ETAP escribe esa corriente en la columna Ia Mag (col 28) de la
        fila "Total". Eso es exactamente el "3I0 Symm Current" de la tabla.
    """
    results = []

    for i, row in df.iterrows():
        cell = str(row[1]).strip()

        if "Fault at bus:" not in cell:
            continue

        bus_name = cell.replace("Fault at bus:", "").strip()
        kv_prefault = extract_prefault_kv(df, i)

        # Buscar la fila "Total" dentro de las próximas 20 filas
        total_row = None
        for j in range(i, min(i + 20, len(df))):
            if str(df.iloc[j][5]).strip() == "Total":
                total_row = j
                break

        if total_row is None:
            print(f"   No se encontró fila 'Total' para bus: {bus_name}")
            continue

        trow = df.iloc[total_row]

        # col 28 = Ia Mag [kA]  → esta es la corriente total 3·I0 en SLG
        ia_kA = float(trow[28])
        # col 42 = I1  (secuencia positiva)
        i1_kA = float(trow[42])

        results.append(
            {
                "Bus name":               bus_name,
                "Bus (kV)":               kv_prefault,
                "3I0 Symm Current 1/2 cycle [kA]": round(ia_kA, 2),
                "I1 [kA]":                round(i1_kA, 4),
            }
        )
        print(f"   ✔ {bus_name:15s}  {kv_prefault:6.2f} kV  "
              f"Ia={ia_kA:.4f} kA  I1={i1_kA:.4f} kA")

    return results


# ──────────────────────────────────────────────────────────────────────────────
# PASO 4 — Exportar al Excel con el formato de la tabla de la imagen
# ──────────────────────────────────────────────────────────────────────────────
def export_excel(
    results: list[dict],
    output_path: str,
    table_title: str = "Table 12. SLG short circuit current results",
    total_mw: float = None,
):
    """
    Genera el .xlsx imitando el estilo de la tabla de la imagen:

        ┌─────────────────────────────────────────────────────┐
        │  Table 12. SLG short circuit current results        │  ← título naranja
        ├──────────────────┬───────────────────────────────────┤
        │ 1-Phase Fault    │ Total Fault Currents – X.X MW     │  ← fila naranja
        ├──────────────────┬───────────┬───────────────────────┤
        │ Bus name         │ Bus (kV)  │ 3I0 Symm Current      │  ← cabecera naranja
        │                  │           │ 1/2 cycle [kA]        │
        ├──────────────────┼───────────┼───────────────────────┤
        │ BUS_A            │   34.5    │         6.71          │
        │ BUS_B            │   34.5    │         6.81          │
        └──────────────────┴───────────┴───────────────────────┘
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SLG Results"

    # ── Paleta de colores (tomada de la imagen) ────────────────────────────
    ORANGE_DARK   = PatternFill("solid", start_color="C55A11", end_color="C55A11")  # título
    ORANGE_MID    = PatternFill("solid", start_color="ED7D31", end_color="ED7D31")  # sub-título
    ORANGE_LIGHT  = PatternFill("solid", start_color="F4B183", end_color="F4B183")  # cabeceras
    ALT_FILL      = PatternFill("solid", start_color="FCE4D6", end_color="FCE4D6")  # filas alternas
    WHITE_FILL    = PatternFill("solid", start_color="FFFFFF", end_color="FFFFFF")  # filas normales

    # ── Fuentes ────────────────────────────────────────────────────────────
    def fnt(bold=False, color="000000", size=10, italic=False):
        return Font(name="Calibri", bold=bold, color=color, size=size, italic=italic)

    # ── Bordes ─────────────────────────────────────────────────────────────
    thin   = Side(style="thin",   color="C55A11")
    medium = Side(style="medium", color="C55A11")
    def brd(l=thin, r=thin, t=thin, b=thin):
        return Border(left=l, right=r, top=t, bottom=b)
    def mbrd():
        return Border(left=medium, right=medium, top=medium, bottom=medium)

    ctr = Alignment(horizontal="center", vertical="center", wrap_text=True)
    lft = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    # ── Fila 1: Título de la tabla ─────────────────────────────────────────
    ws.merge_cells("A1:C1")
    ws["A1"] = table_title
    ws["A1"].font      = fnt(bold=True, color="FFFFFF", size=11)
    ws["A1"].fill      = ORANGE_DARK
    ws["A1"].alignment = ctr
    ws["A1"].border    = mbrd()
    ws.row_dimensions[1].height = 22

    # ── Fila 2: "1-Phase Fault | Total Fault Currents – X.X MW" ───────────
    ws["A2"] = "1-Phase Fault"
    ws["A2"].font      = fnt(bold=True, color="FFFFFF", size=10)
    ws["A2"].fill      = ORANGE_MID
    ws["A2"].alignment = ctr
    ws["A2"].border    = brd()

    mw_text = f"Total Fault Currents – {total_mw} MW" if total_mw else "Total Fault Currents"
    ws.merge_cells("B2:C2")
    ws["B2"] = mw_text
    ws["B2"].font      = fnt(bold=True, color="FFFFFF", size=10)
    ws["B2"].fill      = ORANGE_MID
    ws["B2"].alignment = ctr
    ws["B2"].border    = brd()
    ws.row_dimensions[2].height = 18

    # ── Fila 3: Sub-cabecera "380 Symm Current" ────────────────────────────
    ws["A3"] = ""
    ws["A3"].fill   = ORANGE_LIGHT
    ws["A3"].border = brd()

    ws.merge_cells("B3:C3")
    ws["B3"] = "3I0 Symm Current"
    ws["B3"].font      = fnt(bold=True, size=10)
    ws["B3"].fill      = ORANGE_LIGHT
    ws["B3"].alignment = ctr
    ws["B3"].border    = brd()
    ws.row_dimensions[3].height = 16

    # ── Fila 4: Cabeceras de columnas ──────────────────────────────────────
    col_headers = ["Bus name", "Bus (kV)", "1/2 cycle [kA]"]
    for c, h in enumerate(col_headers, 1):
        cell = ws.cell(row=4, column=c, value=h)
        cell.font      = fnt(bold=True, size=10)
        cell.fill      = ORANGE_LIGHT
        cell.alignment = ctr
        cell.border    = brd()
    ws.row_dimensions[4].height = 18

    # ── Filas de datos ─────────────────────────────────────────────────────
    for idx, rec in enumerate(results):
        row_num = idx + 5
        fill    = ALT_FILL if idx % 2 == 0 else WHITE_FILL

        # Col A: Bus name
        ca = ws.cell(row=row_num, column=1, value=rec["Bus name"])
        ca.font = fnt(size=10)
        ca.fill = fill
        ca.alignment = lft
        ca.border = brd()

        # Col B: Bus (kV)
        cb = ws.cell(row=row_num, column=2, value=rec["Bus (kV)"])
        cb.font = fnt(size=10)
        cb.fill = fill
        cb.alignment = ctr
        cb.border = brd()
        cb.number_format = "0.0"

        # Col C: 3I0 Symm Current
        cc = ws.cell(row=row_num, column=3, value=rec["3I0 Symm Current 1/2 cycle [kA]"])
        cc.font = fnt(size=10)
        cc.fill = fill
        cc.alignment = ctr
        cc.border = brd()
        cc.number_format = "0.00"

        ws.row_dimensions[row_num].height = 16

    # ── Anchos de columna ──────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 18

    wb.save(output_path)
    print(f"\n✅ Archivo guardado: {output_path}")


# ──────────────────────────────────────────────────────────────────────────────
# PUNTO DE ENTRADA
# ──────────────────────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(
        description="Extrae resultados SLG de un reporte ETAP Short-Circuit (.xls/.xlsx)"
    )
    parser.add_argument("input",
        help="Ruta al archivo .xls o .xlsx de ETAP")
    parser.add_argument("--output", "-o", default=None,
        help="Nombre del archivo de salida (default: <input>_SLG.xlsx)")
    parser.add_argument("--title", "-t",
        default="Table 12. SLG short circuit current results",
        help="Título de la tabla en el Excel")
    parser.add_argument("--mw", "-m", type=float, default=None,
        help="Potencia total del sistema en MW para la cabecera (ej: 2.8)")
    args = parser.parse_args()

    if args.output is None:
        p = Path(args.input)
        args.output = str(p.with_name(p.stem + "_SLG.xlsx"))

    print("=" * 55)
    print("  ETAP Short-Circuit — Extractor SLG")
    print("=" * 55)

    df      = read_report(args.input)
    results = extract_slg_results(df)

    if not results:
        print("⚠️  No se encontraron bloques de falla en el archivo.")
        sys.exit(1)

    print(f"\n→ {len(results)} buses de falla encontrados")
    export_excel(results, args.output,
                 table_title=args.title,
                 total_mw=args.mw)


if __name__ == "__main__":
    main()