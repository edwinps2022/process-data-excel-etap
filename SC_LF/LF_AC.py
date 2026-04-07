"""
ETAP Cable Summary — Extractor TB + LBD
========================================

Combina dos archivos ETAP para generar una tabla resumen con columnas:
    TB: Bus name | Size | Active Power | Ampacity | Current | Active Losses | Vd
    LBD: Bus name | Size | Active Power | Ampacity | Current | Active Losses | Vd
    Vd total (suma TB + LBD)

Fuentes de datos:
    - Branch Flow Summary (.xls): Active Power, Current, Active Losses, Vd
    - cable_data (.xls) pestaña Ampacity: Size (col G) y Ampacity (col U)

Uso desde la terminal:
    python etap_cable_summary.py
    python etap_cable_summary.py --branch mi_branch.xls --cable mi_cable.xls
    python etap_cable_summary.py --branch mi_branch.xls --cable mi_cable.xls --output resultado.xlsx
    python etap_cable_summary.py --tb_prefix "TB" --lbd_prefix "LBD"

Requisitos (instalar una sola vez):
    pip install pandas xlrd openpyxl
"""

import argparse
import sys
from pathlib import Path

import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# ──────────────────────────────────────────────────────────────────────────────
# PASO 1 — Leer Branch Flow Summary
# ──────────────────────────────────────────────────────────────────────────────
def read_branch_flow(filepath: str) -> dict:
    """
    Lee el archivo Branch Flow Summary de ETAP.

    Columnas usadas (índice base 0):
        col  1 → ID del cable (nombre)
        col  2 → Type ("Cable" es el único que nos interesa)
        col 16 → Active Power To-From [kW]  (potencia entregada)
        col 21 → Active Losses [kW]
        col 24 → Current [Amp]
        col 33 → Vd [%]

    Retorna un diccionario: { "LBD 1.1.1.1": { active_power, losses, current, vd }, ... }
    """
    path = Path(filepath)
    if not path.exists():
        sys.exit(f"❌ Archivo no encontrado: {filepath}")

    engine = "xlrd" if path.suffix.lower() == ".xls" else "openpyxl"
    print(f"📂 Leyendo Branch Flow: {path.name}")
    df = pd.read_excel(filepath, engine=engine, header=None)

    data = {}
    for i in range(len(df)):
        row = df.iloc[i]
        name  = str(row[1]).strip()
        btype = str(row[2]).strip()

        if btype != "Cable" or not name or name == "nan":
            continue
        try:
            data[name] = {
                "active_power": abs(float(row[16])),
                "losses":       abs(float(row[21])),
                "current":      abs(float(row[24])),
                "vd":           abs(float(row[33])),
            }
        except (ValueError, TypeError):
            pass

    print(f"   → {len(data)} cables cargados")
    return data


# ──────────────────────────────────────────────────────────────────────────────
# PASO 2 — Leer cable_data pestaña Ampacity
# ──────────────────────────────────────────────────────────────────────────────
def read_cable_data(filepath: str, sheet: str = "Ampacity") -> dict:
    """
    Lee la pestaña Ampacity del archivo cable_data.

    Columnas usadas (índice base 0):
        col  1 → ID del cable (nombre)
        col  6 → Size [AWG/kcmil]        — columna G en Excel
        col 20 → Derated Amp (Ampacity)  — columna U en Excel

    Datos empiezan en la fila 6 (0-based) porque las primeras filas
    son títulos y sub-cabeceras del reporte ETAP.

    Retorna un diccionario: { "LBD 1.1.1.1": { size, ampacity }, ... }

    ¿Por qué col G y col U?
        Col G (index 6) = "Size AWG/kcmil" → calibre del conductor
        Col U (index 20) = "Derated Amp"   → amperaje considerando
                           todos los factores de corrección aplicados
                           (temperatura, agrupamiento, etc.)
    """
    path = Path(filepath)
    if not path.exists():
        sys.exit(f"❌ Archivo no encontrado: {filepath}")

    engine = "xlrd" if path.suffix.lower() == ".xls" else "openpyxl"
    print(f"📂 Leyendo cable_data ({sheet}): {path.name}")

    df = pd.read_excel(filepath, engine=engine, sheet_name=sheet, header=None)

    data = {}
    for i in range(6, len(df)):          # datos reales desde fila 6
        row  = df.iloc[i]
        name = str(row[1]).strip()
        if not name or name == "nan":
            continue
        try:
            amp = float(row[20])
        except (ValueError, TypeError):
            amp = 0.0
        data[name] = {
            "size":     str(row[6]).strip(),
            "ampacity": amp,
        }

    print(f"   → {len(data)} cables cargados")
    return data


# ──────────────────────────────────────────────────────────────────────────────
# PASO 3 — Emparejar TB ↔ LBD por sufijo numérico
# ──────────────────────────────────────────────────────────────────────────────
def build_pairs(branch_data: dict,
                tb_prefix: str = "TB",
                lbd_prefix: str = "LBD") -> list:
    """
    Empareja cables TB y LBD que comparten el mismo sufijo numérico.

    Ejemplo:
        "TB 1.1.1.1"  ↔  "LBD 1.1.1.1"   → suffix = "1.1.1.1"
        "TB 1.2.3.4"  ↔  "LBD 1.2.3.4"   → suffix = "1.2.3.4"

    El emparejamiento es automático: busca los nombres que empiezan con
    tb_prefix y lbd_prefix, extrae el sufijo (todo lo que va después del
    prefijo + espacio) y los agrupa.

    Retorna lista de tuplas: [ ("1.1.1.1", {"tb": "TB 1.1.1.1", "lbd": "LBD 1.1.1.1"}), ... ]
    ordenada numéricamente por sufijo.

    ¿Cómo ajustar para otros proyectos?
        Si los cables se llaman distinto, cambia tb_prefix y lbd_prefix:
        --tb_prefix "TRUNK" --lbd_prefix "LOAD"
        El script buscará "TRUNK x.x.x" ↔ "LOAD x.x.x"
    """
    pairs = {}
    prefix_tb  = tb_prefix  + " "
    prefix_lbd = lbd_prefix + " "

    for name in branch_data:
        if name.startswith(prefix_lbd):
            suffix = name[len(prefix_lbd):]
            pairs.setdefault(suffix, {})["lbd"] = name
        elif name.startswith(prefix_tb):
            suffix = name[len(prefix_tb):]
            pairs.setdefault(suffix, {})["tb"] = name

    def sort_key(s):
        try:
            return [int(x) for x in s.split(".")]
        except ValueError:
            return [0]

    sorted_pairs = sorted(pairs.items(), key=lambda x: sort_key(x[0]))
    print(f"   → {len(sorted_pairs)} pares TB+LBD encontrados")
    return sorted_pairs


# ──────────────────────────────────────────────────────────────────────────────
# PASO 4 — Exportar a Excel con formato profesional
# ──────────────────────────────────────────────────────────────────────────────
def export_excel(sorted_pairs: list,
                 branch_data: dict,
                 cable_info: dict,
                 output_path: str):
    """
    Genera el archivo .xlsx con la tabla doble TB | LBD | Vd total.

    Estructura:
        Fila 1: Título principal
        Fila 2: Info del proyecto
        Fila 3: Etiquetas de sección (TB / LBD / Vd total)
        Fila 4: Cabeceras de columnas
        Filas 5+: Datos (una fila por par TB+LBD)
        Última fila: Totales con fórmulas Excel
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cable Summary"

    # ── Paleta ──────────────────────────────────────────────────────────────
    BLUE_DARK  = PatternFill("solid", start_color="1F3864", end_color="1F3864")
    BLUE_MID   = PatternFill("solid", start_color="2E75B6", end_color="2E75B6")
    TEAL_HDR   = PatternFill("solid", start_color="006B6B", end_color="006B6B")
    GREEN_HDR  = PatternFill("solid", start_color="375623", end_color="375623")
    TEAL_ALT   = PatternFill("solid", start_color="D6EAF8", end_color="D6EAF8")
    GREEN_ALT  = PatternFill("solid", start_color="E8F5E9", end_color="E8F5E9")
    WHITE_FILL = PatternFill("solid", start_color="FFFFFF", end_color="FFFFFF")
    TOTAL_FILL = PatternFill("solid", start_color="FFF2CC", end_color="FFF2CC")

    thin   = Side(style="thin",   color="B0BEC5")
    medium = Side(style="medium", color="1F3864")

    def brd():
        return Border(left=thin, right=thin, top=thin, bottom=thin)

    def mbrd():
        return Border(left=medium, right=medium, top=medium, bottom=medium)

    ctr = Alignment(horizontal="center", vertical="center", wrap_text=True)
    lft = Alignment(horizontal="left",   vertical="center")

    def fnt(bold=False, color="000000", size=9, italic=False):
        return Font(name="Arial", bold=bold, color=color, size=size, italic=italic)

    # Helpers para sacar valores con default seguro
    def get_d(name, key, default=0.0):
        return branch_data.get(name, {}).get(key, default)

    def get_c(name, key, default=""):
        return cable_info.get(name, {}).get(key, default)

    # ── Fila 1: Título ───────────────────────────────────────────────────────
    ws.merge_cells("A1:O1")
    ws["A1"] = "Cable Summary Report — Branch Flow (TB) & Load Bus (LBD)"
    ws["A1"].font      = fnt(bold=True, color="FFFFFF", size=12)
    ws["A1"].fill      = BLUE_DARK
    ws["A1"].alignment = ctr
    ws.row_dimensions[1].height = 26

    # ── Fila 2: Info del proyecto ────────────────────────────────────────────
    ws.merge_cells("A2:O2")
    ws["A2"] = "Project: Franklyn  |  Study Case: LF DC  |  Config.: Normal  |  Date: 04-06-2026"
    ws["A2"].font      = fnt(italic=True, size=9, color="555555")
    ws["A2"].alignment = ctr
    ws.row_dimensions[2].height = 14

    # ── Fila 3: Secciones ────────────────────────────────────────────────────
    ws.merge_cells("A3:G3")
    ws["A3"] = "TB — Trunkline Bus"
    ws["A3"].font = fnt(bold=True, color="FFFFFF", size=10)
    ws["A3"].fill = TEAL_HDR
    ws["A3"].alignment = ctr

    ws.merge_cells("H3:N3")
    ws["H3"] = "LBD — Load Bus"
    ws["H3"].font = fnt(bold=True, color="FFFFFF", size=10)
    ws["H3"].fill = GREEN_HDR
    ws["H3"].alignment = ctr

    ws["O3"] = "Vd total\nsuma TB+LBD"
    ws["O3"].font = fnt(bold=True, color="FFFFFF", size=9)
    ws["O3"].fill = BLUE_MID
    ws["O3"].alignment = ctr
    ws.row_dimensions[3].height = 28

    # ── Fila 4: Cabeceras de columnas ────────────────────────────────────────
    headers = [
        "Bus name", "Size\n(AWG)", "Active Power\n(kW)", "Ampacity\n(A)",
        "Current\n(A)", "Active Losses\n(kW)", "Vd\n(%)",
        "Bus name", "Size\n(AWG)", "Active Power\n(kW)", "Ampacity\n(A)",
        "Current\n(A)", "Active Losses\n(kW)", "Vd\n(%)",
        "Vd total\n(%)"
    ]
    fills_hdr = [TEAL_HDR] * 7 + [GREEN_HDR] * 7 + [BLUE_MID]
    for c, (h, f) in enumerate(zip(headers, fills_hdr), 1):
        cell = ws.cell(row=4, column=c, value=h)
        cell.font      = fnt(bold=True, color="FFFFFF", size=9)
        cell.fill      = f
        cell.alignment = ctr
        cell.border    = brd()
    ws.row_dimensions[4].height = 32

    # ── Filas de datos ───────────────────────────────────────────────────────
    FMT_NUM = "#,##0.000"

    for idx, (suffix, p) in enumerate(sorted_pairs):
        row_num  = idx + 5
        tb_name  = p.get("tb",  "")
        lbd_name = p.get("lbd", "")
        tb_fill  = TEAL_ALT  if idx % 2 == 0 else WHITE_FILL
        lbd_fill = GREEN_ALT if idx % 2 == 0 else WHITE_FILL

        # ── Columnas TB (A–G = 1–7) ─────────────────────────────────────────
        tb_vals = [
            tb_name,
            get_c(tb_name, "size"),
            get_d(tb_name, "active_power"),
            get_c(tb_name, "ampacity"),
            get_d(tb_name, "current"),
            get_d(tb_name, "losses"),
            get_d(tb_name, "vd"),
        ]
        for c, val in enumerate(tb_vals, 1):
            cell           = ws.cell(row=row_num, column=c, value=val)
            cell.font      = fnt(size=9)
            cell.fill      = tb_fill
            cell.border    = brd()
            cell.alignment = lft if c == 1 else ctr
            if c in (3, 4, 5, 6, 7):
                cell.number_format = FMT_NUM

        # ── Columnas LBD (H–N = 8–14) ───────────────────────────────────────
        lbd_vals = [
            lbd_name,
            get_c(lbd_name, "size"),
            get_d(lbd_name, "active_power"),
            get_c(lbd_name, "ampacity"),
            get_d(lbd_name, "current"),
            get_d(lbd_name, "losses"),
            get_d(lbd_name, "vd"),
        ]
        for c, val in enumerate(lbd_vals, 8):
            cell           = ws.cell(row=row_num, column=c, value=val)
            cell.font      = fnt(size=9)
            cell.fill      = lbd_fill
            cell.border    = brd()
            cell.alignment = lft if c == 8 else ctr
            if c in (10, 11, 12, 13, 14):
                cell.number_format = FMT_NUM

        # ── Columna Vd total (O = 15) ────────────────────────────────────────
        vd_total = get_d(tb_name, "vd") + get_d(lbd_name, "vd")
        ct = ws.cell(row=row_num, column=15, value=vd_total)
        ct.font           = fnt(bold=True, size=9)
        ct.fill           = TOTAL_FILL
        ct.border         = brd()
        ct.alignment      = ctr
        ct.number_format  = FMT_NUM

        ws.row_dimensions[row_num].height = 15

    # ── Fila de totales ──────────────────────────────────────────────────────
    n         = len(sorted_pairs)
    last_data = 4 + n
    tot_row   = last_data + 1

    ws.merge_cells(f"A{tot_row}:B{tot_row}")
    ws[f"A{tot_row}"] = f"TOTALS  ({n} cables)"
    ws[f"A{tot_row}"].font      = fnt(bold=True, size=9, color="FFFFFF")
    ws[f"A{tot_row}"].fill      = BLUE_DARK
    ws[f"A{tot_row}"].alignment = lft
    ws[f"A{tot_row}"].border    = mbrd()

    # Fórmulas SUM para: Active Power TB(C), Losses TB(F),
    #                    Active Power LBD(J), Losses LBD(M), Vd total(O)
    sum_cols = [(3, "C"), (6, "F"), (10, "J"), (13, "M"), (15, "O")]
    for col_num, col_letter in sum_cols:
        cell = ws.cell(row=tot_row, column=col_num,
                       value=f"=SUM({col_letter}5:{col_letter}{last_data})")
        cell.font          = fnt(bold=True, size=9, color="FFFFFF")
        cell.fill          = BLUE_DARK
        cell.alignment     = ctr
        cell.border        = mbrd()
        cell.number_format = FMT_NUM

    for c in range(1, 16):
        cell = ws.cell(row=tot_row, column=c)
        if cell.value is None:
            cell.fill   = BLUE_DARK
            cell.border = mbrd()
    ws.row_dimensions[tot_row].height = 18

    # ── Anchos de columna ────────────────────────────────────────────────────
    col_widths = [22, 8, 14, 11, 11, 14, 10,
                  22, 8, 14, 11, 11, 14, 10, 11]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Borde grueso entre secciones TB | LBD y LBD | Vd total
    for r in range(3, tot_row + 1):
        ws.cell(row=r, column=7).border = Border(
            left=Side(style="thin", color="B0BEC5"),
            right=Side(style="medium", color="1F3864"),
            top=Side(style="thin", color="B0BEC5"),
            bottom=Side(style="thin", color="B0BEC5"))
        ws.cell(row=r, column=15).border = Border(
            left=Side(style="medium", color="1F3864"),
            right=Side(style="thin", color="B0BEC5"),
            top=Side(style="thin", color="B0BEC5"),
            bottom=Side(style="thin", color="B0BEC5"))

    ws.freeze_panes = "A5"
    wb.save(output_path)
    print(f"✅ Archivo guardado: {output_path}")


# ──────────────────────────────────────────────────────────────────────────────
# PUNTO DE ENTRADA
# ──────────────────────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(
        description="Genera tabla Cable Summary (TB + LBD) desde archivos ETAP"
    )
    parser.add_argument(
        "--branch", "-b",
        default="LF_DC_DL1S_-_Branch_Flow_Summary.xls",
        help="Ruta al Branch Flow Summary de ETAP (.xls o .xlsx)"
    )
    parser.add_argument(
        "--cable", "-c",
        default="cable_data.xls",
        help="Ruta al archivo cable_data (.xls o .xlsx)"
    )
    parser.add_argument(
        "--output", "-o",
        default=None,
        help="Nombre del archivo de salida (default: Cable_Summary.xlsx)"
    )
    parser.add_argument(
        "--tb_prefix",
        default="TB",
        help="Prefijo de los cables Trunkline (default: 'TB')"
    )
    parser.add_argument(
        "--lbd_prefix",
        default="LBD",
        help="Prefijo de los cables Load Bus (default: 'LBD')"
    )
    parser.add_argument(
        "--ampacity_sheet",
        default="Ampacity",
        help="Nombre de la pestaña de amperaje en cable_data (default: 'Ampacity')"
    )
    args = parser.parse_args()

    if args.output is None:
        args.output = "Cable_Summary.xlsx"

    print("=" * 55)
    print("  ETAP Cable Summary — Extractor TB + LBD")
    print("=" * 55)

    branch_data = read_branch_flow(args.branch)
    cable_info  = read_cable_data(args.cable, sheet=args.ampacity_sheet)
    sorted_pairs = build_pairs(branch_data,
                               tb_prefix=args.tb_prefix,
                               lbd_prefix=args.lbd_prefix)

    if not sorted_pairs:
        print("⚠️  No se encontraron pares TB+LBD.")
        print("   Verifica --tb_prefix y --lbd_prefix.")
        sys.exit(1)

    export_excel(sorted_pairs, branch_data, cable_info, args.output)


if __name__ == "__main__":
    main()